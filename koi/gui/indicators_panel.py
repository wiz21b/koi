from datetime import date

import openpyxl
from PySide.QtCore import Qt, Slot, QObject, Signal
from PySide.QtGui import QVBoxLayout, QHBoxLayout, QPushButton
from openpyxl.styles import Font

# This fix was disabled...
# from koi.tools.fix_openpyxl import ignore_openpyxl_constants
# ignore_openpyxl_constants()

from koi.base_logging import mainlog
from koi.charts.chart_widget import ChartWidget
from koi.dao import dao
from koi.date_utils import month_before, month_after
from koi.gui.dialog_utils import KPIView, SubFrame, make_progress, NavBar, TitleWidget
from koi.gui.horse_panel import HorsePanel
from koi.reporting.utils import make_temp_file, open_xlsx
from koi.translators import date_to_my

class IndicatorsPanel(HorsePanel):

    MONTHS_STEP = 1
    """ number of month to go forward/backward when choosing the period of indicators """

    MONTHS_PERIOD = 6

    def _indicators_run(self, indicators):
        all_indictors = []

        for line in indicators:
            for ind in line:
                if isinstance(ind, KPIView) or isinstance(ind, ChartWidget):
                    all_indictors.append(ind)
                elif type(ind) == list:
                    all_indictors += self._indicators_run(ind)

        return all_indictors

    def _layout_indicators(self, indicators):
        subframes = []
        ind_vlayout = QVBoxLayout()

        for line in indicators:
            hboxlayout = QHBoxLayout()
            hboxlayout.addStretch()
            hboxlayout.setAlignment( Qt.AlignTop)

            for ind in line[1:]:

                if type(ind) == list:
                    deeper_layout, deeper_subframes = self._layout_indicators(ind)
                    subframes += deeper_subframes
                    hboxlayout.addLayout( deeper_layout )
                else:
                    hboxlayout.addWidget(ind)

            hboxlayout.addStretch()
            s = SubFrame( line[0],hboxlayout,self)
            subframes.append(s)

            ind_vlayout.addWidget(s)

        return ind_vlayout, subframes

    @Slot()
    def _clear_cache(self):
        self.remote_indicators_service.clear_caches()
        self.refresh_action()

    @Slot(date)
    def _month_changed(self, d):
        self.refresh_action()

    @Slot()
    def refresh_action(self): # Reimplements parent method !
        self.remote_indicators_service.clear_caches()
        mainlog.debug("Refreshing {}".format(self.month_chooser.base_date))
        begin, end = month_before(self.month_chooser.base_date, self.MONTHS_PERIOD-1), self.month_chooser.base_date

        progress = make_progress(_("Collecting data..."), sum([len(x) - 1 for x in self.indicators]))

        mainlog.debug("Indicator run = {}".format( self._indicators_run( self.indicators)))
        for ind in self._indicators_run( self.indicators):
            ind._gather_data(begin, end)
            progress.setValue(progress.value() + 1)
            ind.repaint()

        for sf in self._sub_frames:
            if sf.original_title():
                sf.set_title(
                    sf.original_title().replace( "%MONTH%",
                                                 date_to_my( self.month_chooser.base_date, full=True)))

    @Slot()
    def _export_to_excel(self):


        workbook = openpyxl.Workbook()
        sheet = workbook.active

        progress = make_progress(_("Collecting data..."), sum([len(x) - 1 for x in  self.indicators]))

        row = 1
        for ind in self._indicators_run(self.indicators):
            chart = ind.chart
            data = ind.chart.original_data

            #print(chart.x_legends[serie_ndx])

            sheet.cell(row=row, column=1, value=chart.title)
            sheet.cell(row=row, column=1).font = Font(bold=True)

            row += 1
            for col in range(len(chart.x_legends)):
                sheet.cell(row=row, column=2 + col, value=chart.x_legends[col])

            row += 1
            for serie_ndx in range(len(data)):
                sheet.cell(row=row, column=1, value=chart.legends[serie_ndx])
                for col in range(len(data[serie_ndx])):
                    sheet.cell(row=row, column=2+col, value=data[serie_ndx][col])
                row += 1

            row += 2 # Space between kpi's
            progress.setValue( progress.value() + 1)

        name = make_temp_file('indicators_','xlsx')
        workbook.save(name)
        open_xlsx(name)


    def __init__(self, parent, remote_indicators_service, title, indicators):
        super(IndicatorsPanel, self).__init__(parent)

        self.remote_indicators_service = remote_indicators_service

        self.month_chooser = MonthChooser( self.MONTHS_STEP)
        month_before, month_today, month_after = self.month_chooser.navbar_buttons_tuples()
        self.month_chooser.month_changed.connect(self._month_changed)
        self.set_panel_title( title)

        self.indicators = indicators

        vlayout = QVBoxLayout(self)

        navigation = NavBar(None,
                            [month_before,
                             month_today,
                             month_after,
                             (_("Recompute"), self._clear_cache),
                             (_("Export to Excel"), self._export_to_excel)])

        self.title_box = TitleWidget(title, self, navigation)  # + date.today().strftime("%B %Y"),self)
        vlayout.addWidget(self.title_box)

        # Pay attention, subframes holds a reference to each
        # subframe widget !

        ind_layout, self._sub_frames = self._layout_indicators(self.indicators)
        vlayout.addLayout( ind_layout)
        self.setLayout(vlayout)



class MonthChooser(QObject):
    month_changed = Signal(date)

    def __init__(self, step_size_in_months = 1):
        super(MonthChooser, self).__init__()
        self.base_date = date.today()
        self.step_size = step_size_in_months # month
        self.minimum_date = dao.order_dao.get_oldest_order_creation_date()

    def _activate_buttons(self):
        self._prev_button.setEnabled( self._month_in_range( month_before(self.base_date, self.step_size)))
        self._next_button.setEnabled( self._month_in_range( month_after(self.base_date, self.step_size)))

    def _set_month(self, d):
        self.base_date = d
        self._activate_buttons()
        self.month_changed.emit(self.base_date)

    def navbar_buttons_tuples(self):
        # FIXME Ownership issues !
        # 1. we create those buttons
        # 2. they're added to a navbar layout. That layout takes ownership
        # 3. The layout is destroyed at some point
        # 4. So these refrences to the buttons must be destroyed as well
        # => risk of double destruction :-(

        self._prev_button = QPushButton(_("{} months before").format(self.step_size))
        self._today_button = QPushButton(_("Today"))
        self._next_button = QPushButton(_("{} months after").format(self.step_size))
        self._activate_buttons()

        return (self._prev_button, self.month_before),  (self._today_button,self.month_today), (self._next_button, self.month_after)

    def _month_in_range(self, m : date):
        """ Month are expected to be within a range. This range
        goes from minus infinity to today.

        :param m: date
        :return: True if m is in the range
        """

        year_today = date.today().year

        if (m.year < year_today or (m.year == year_today and m.month <= date.today().month)) and \
           (m.year > self.minimum_date.year or (m.year == self.minimum_date.year and m.month >= self.minimum_date.month)):
            return True
        else:
            return False



    @Slot()
    def month_today(self):
        self._set_month(date.today())

    @Slot()
    def month_before(self):
        self._set_month(month_before(self.base_date, self.step_size))

    @Slot()
    def month_after(self):
        m = month_after(self.base_date, self.step_size)
        if self._month_in_range(m):
            self._set_month(m)

# class ChartDialog(QDialog):
#     def __init__(self,parent,chart):
#         super(ChartDialog,self).__init__(parent)
#
#         self.buttons = QDialogButtonBox()
#         self.buttons.addButton( QDialogButtonBox.Ok)
#
#         vlayout = QVBoxLayout()
#         vlayout.addWidget(chart)
#         vlayout.addWidget(self.buttons)
#         self.buttons.accepted.connect(self.accept)
#
#         self.setMinimumSize(800,600)
#
#         self.setLayout(vlayout)
#
#
#     @Slot()
#     def accept(self):
#         super(ChartDialog,self).accept()

